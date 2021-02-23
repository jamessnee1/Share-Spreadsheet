import os
import sys
import time
import datetime
import yfinance as yf
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

#Format: Ticker Symbol, Number of shares
AUS_TICKERS = {
"ANZ.AX" : 48,
"APA.AX" : 45,
"APT.AX" : 76,
"ARF.AX" : 230,
"BOQ.AX" : 77,
"BSL.AX" : 188,
"CBA.AX" : 29,
"DTS.AX" : 3333,
"FMG.AX" : 200,
"GEM.AX" : 500,
"PL8.AX" : 1351,
"RMD.AX" : 20,
"VHY.AX" : 9,
"Z1P.AX" : 262,
"AGNC" : 3,
"AMC" : 1,
"GAIN" : 6,
"GPRO" : 1,
"ORC" : 13,
"PSEC" : 10,
"STAG" : 1
}

def writeTitle(ws, title, width, column):
    cell = column + "1"
    ws.column_dimensions[column].width = width
    ws[cell] = title
    ws[cell].font = Font(bold=True)
    

def writeData(ws, row, column, value):
    ws.cell(row=row, column=column, value=str(value))
    

#Debugging
#print(AUS_TICKERS)

start = datetime.datetime.now()

#Create excel
wb = Workbook()  
ws = wb.active
strDate = start.strftime("%d_%m_%Y")
dest_filename = 'SharePortfolio_' + strDate + '.xlsx'

#If file exists, delete it
if os.path.exists(dest_filename):
    os.remove(dest_filename)
    
ws.title = "Shares " + start.strftime("%d_%m_%Y")
#Write excel column headers
writeTitle(ws, "Company Name", 55, "A")
writeTitle(ws, "Ticker Symbol", 12, "B")
writeTitle(ws, "Sector", 25, "C")
writeTitle(ws, "Current Market Price (ask)", 25, "D")
writeTitle(ws, "Number of shares", 25, "E")
writeTitle(ws, "Total share value", 25, "F")
writeTitle(ws, "Last dividend per share", 25, "G")
writeTitle(ws, "Dividend yield", 25, "H")
writeTitle(ws, "Ex-dividend date", 25, "I")
writeTitle(ws, "Estimated Income", 25, "J")
    
print("Downloading stock data from Yahoo Finance...")
rowCount = 2

totalDividendIncome = 0.00
totalNumOfShares = 0
totalPortfolioValue = 0.00

for company in AUS_TICKERS:
    print("Getting data for " + company)
    ticker = yf.Ticker(company)
    #print("next event: " + str(ticker.calendar))
    #print(ticker.dividends)
    
    try:
        sector = str(ticker.info['sector'])
    except KeyError: sector = "N/A"
    
    longName = str(ticker.info['longName'])
    lastDividendValue = str(ticker.info['lastDividendValue'])
    marketPrice = str(ticker.info['ask'])
    dividendYield = str(ticker.info['dividendYield'])
    
    if dividendYield == 'None':
        dividendYield = 0.00
    
    if lastDividendValue == 'None':
        lastDividendValue = 0.00
        
        
    estimatedIncome = round(float(AUS_TICKERS[company]) * float(lastDividendValue), 3)
    totalDividendIncome = round(totalDividendIncome + estimatedIncome, 3)
    
    totalShareValue = round(float(AUS_TICKERS[company]) * float(marketPrice), 3)
    totalNumOfShares = round(totalNumOfShares + AUS_TICKERS[company])
    totalPortfolioValue = round(totalPortfolioValue + totalShareValue, 3)
    
    try:
        exDate = ticker.calendar['Value'][0]
    except KeyError: exDate = "N/A"
    except TypeError: exDate = "N/A"
    
    #Data from API
    writeData(ws, rowCount, 1, longName)
    writeData(ws, rowCount, 2, company)
    writeData(ws, rowCount, 3, sector)
    writeData(ws, rowCount, 4, "$" + str(marketPrice))
    writeData(ws, rowCount, 5, str(AUS_TICKERS[company]))
    writeData(ws, rowCount, 6, "$" + str(totalShareValue))
    writeData(ws, rowCount, 7, "$" + str(lastDividendValue))
    writeData(ws, rowCount, 8, str(dividendYield))
    writeData(ws, rowCount, 9, str(exDate))
    writeData(ws, rowCount, 10, "$" + str(estimatedIncome))
    rowCount = rowCount + 1

#Write totals to spreadsheet
writeData(ws, rowCount, 1, "TOTAL:")
writeData(ws, rowCount, 5, totalNumOfShares)
writeData(ws, rowCount, 6, "$" + str(totalPortfolioValue))
writeData(ws, rowCount, 10, "$" + str(totalDividendIncome))

#Save spreadsheet
wb.save(filename = dest_filename)

end = datetime.datetime.now()
elapsed = end - start

print("Script took " + str(elapsed) + " to execute")